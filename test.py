from openpyxl import load_workbook
import boto3

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']


def change(emotion):
    current = ['ANGRY', 'CONFUSED', 'DISGUSTED', 'HAPPY', 'CALM', 'SAD', 'SURPRISED']
    transform = ['ANGRY', 'CONFUSED', 'DISGUST', 'HAPPY', 'NEUTRAL', 'SAD', 'SURPRISED']
    result = ''
    for i in range(8):
        if current[i] == emotion:
            result = transform[i]
            break
    return result


imageFile = 'images/GW34/GW34-A/046.jpg'
label = 'SAD'
row=123

if __name__ == "__main__":
    bucket = 'bucket'
    client = boto3.client('rekognition')

    with open(imageFile, 'rb') as image:
        response = client.detect_faces(Image={'Bytes': image.read()}, Attributes=['ALL'])
        print(response['FaceDetails'])

    if response['FaceDetails'] == []:
        ws['A' + str(row)] = imageFile
        wb.save('results.xlsx')
    else:
        ANGRY = 0.000
        CONFUSED = 0.000
        DISGUSTED = 0.000
        HAPPY = 0.000
        CALM = 0.000
        SAD = 0.000
        SURPRISED = 0.000

        con = 0.000
        ty = ''
        for emo in response['FaceDetails'][0]['Emotions']:
            confidence = emo['Confidence']
            type = emo['Type']
            if confidence > con:
                con = confidence
                ty = change(type)

        ws['A' + str(row)] = imageFile
        ws['B' + str(row)] = ty
        ws['C' + str(row)] = con
        ws['D' + str(row)] = ty == label
        wb.save('results.xlsx')




