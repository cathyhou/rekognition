from openpyxl import load_workbook

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']

count = 0

for i in range(128):
    if ws['D'+str(i+2)].value:
        count = count + 1

print(count/128.0)