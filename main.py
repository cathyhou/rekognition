import boto3
from openpyxl import load_workbook
import pandas as pd

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']
row = 2


def change(emotion):
    current = ['ANGRY', 'CONFUSED', 'DISGUSTED', 'HAPPY', 'CALM', 'SAD', 'SURPRISED']
    transform = ['ANGRY', 'CONFUSED', 'DISGUST', 'HAPPY', 'NEUTRAL', 'SAD', 'SURPRISED']
    result = ''
    for i in range(8):
        if current[i] == emotion:
            result = transform[i]
            break
    return result


export = pd.read_excel('export.xlsx')

for f in range(len(export)):
    if export['type'].at[f] == 'TEST': # read export file to determine if image is test image, if not move on
        label = export['label'].at[f] # label is actual emotion

        if __name__ == "__main__":
            imageFile = export['name'].at[f][60:] # get rid of extra part of file name
            bucket = 'bucket'
            client = boto3.client('rekognition')

            with open(imageFile, 'rb') as image:
                response = client.detect_faces(Image={'Bytes': image.read()}, Attributes=['ALL']) # run rekognition api
                print(response['FaceDetails'])

            if response['FaceDetails']==[]: # if no face detected, leave row blank+run with original image later
                ws['A' + str(row)] = imageFile
                wb.save('results.xlsx')
                row = row + 1
                print(row)
            else :
                ANGRY = 0.000
                CONFUSED = 0.000
                DISGUSTED = 0.000
                HAPPY = 0.000
                CALM = 0.000
                SAD = 0.000
                SURPRISED = 0.000

                con = 0.000
                ty = ''
                for emo in response['FaceDetails'][0]['Emotions']: # determine emotion with highest confidence
                    confidence = emo['Confidence']
                    type = emo['Type']
                    if confidence > con:
                        con = confidence
                        ty = change(type)

                ws['A'+str(row)] = imageFile
                ws['B' + str(row)] = ty
                ws['C' + str(row)] = con
                ws['D' + str(row)] = ty==label # boolean (for accuracy)
                wb.save('results.xlsx') # save image file + emotion
                row = row+1
                print(row)


