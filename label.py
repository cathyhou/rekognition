from openpyxl import load_workbook

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']


def toEmotion(emo):
    current = ['AN', 'SC', 'HA', 'NE', 'SA']
    list = ['ANGRY', 'SCARED', 'HAPPY', 'NEUTRAL', 'SAD']
    for i in range(5):
        if emo == current[i]:
            return list[i]
    return 'error'


# determine what correct emotion is from image file path

for i in range(128):
    row = str(i+2)
    img_url = ws['A'+row].value
    emo = img_url[9:11]
    ws['E'+row] = toEmotion(emo)
    wb.save('results.xlsx')
