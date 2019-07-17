import seaborn as sn
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']
array = [[0,0,0,0,0], [0,0,0,0,0], [0,0,0,0,0], [0,0,0,0,0], [0,0,0,0,0]]

# build confusion matrix for rekognition results and automl


def change(emo):
    emotions = ['HAPPY', 'NEUTRAL', 'SCARED', 'ANGRY', 'SAD']
    for i in range(5):
        if emo == emotions[i]:
            return i
    return 'error'


for row in range(128):
    r = str(row + 2)
    predicted = change(ws['B'+r].value)
    true = change(ws['E'+r].value)
    if predicted == 'error' or true == 'error':
        continue
    else:
        current = array[true][predicted]
        array[true][predicted] = current + 1


df_cm = pd.DataFrame(array, columns=["HAPPY","NEUTRAL","SCARED","ANGRY","SAD"], index=["HAPPY","NEUTRAL","SCARED","ANGRY","SAD"])
plt.figure(figsize = (10,10))
sn.set(font_scale=1)
ax = sn.heatmap(df_cm, annot=True,annot_kws={"size": 24}, cbar=False, cmap='Blues', fmt='g')# font size
ax.xaxis.tick_top()
ax.xaxis.set_ticks_position('none')
plt.yticks(va='center', size=16)
plt.xticks(va='center', size=16)
plt.ylabel('True \n', fontsize=20, fontweight='bold')
plt.title('Predicted \n', fontsize=20, fontweight='bold')
plt.xlabel('\n REKOGNITION', fontsize=28, fontweight='heavy')

array2 = array

for row in range(5):
    s = 0
    for col in range(5):
        s = s + array2[row][col]
    for col in range(5):
        current = array2[row][col]
        array2[row][col] = round(current/s * 100.00, 1)

df_cm = pd.DataFrame(array2, columns=["HAPPY","NEUTRAL","SCARED","ANGRY","SAD"], index=["HAPPY","NEUTRAL","SCARED","ANGRY","SAD"])
plt.figure(figsize = (10,10))
sn.set(font_scale=1)
ax = sn.heatmap(df_cm, annot=True,annot_kws={"size": 24}, cbar=False, cmap='Blues', fmt='g')# font size
ax.xaxis.tick_top()
ax.xaxis.set_ticks_position('none')
plt.yticks(va='center', size=16)
plt.xticks(va='center', size=16)
plt.ylabel('True \n', fontsize=20, fontweight='bold')
plt.title('Predicted \n', fontsize=20, fontweight='bold')
plt.xlabel('\n REKOGNITION', fontsize=28, fontweight='heavy')
plt.savefig('RekognitionConfusion.eps')


array3=[[23, 7, 0, 0, 0], [4, 54, 0, 0, 0], [0, 0, 14, 2, 0], [0, 0, 0, 10, 0], [0, 0, 0, 0, 15]]
df_cm = pd.DataFrame(array3, columns=["HAPPY","NEUTRAL","SCARED","ANGRY","SAD"], index=["HAPPY","NEUTRAL","SCARED","ANGRY","SAD"] )
plt.figure(figsize = (10,10))
sn.set(font_scale=1)
ax = sn.heatmap(df_cm, annot=True,annot_kws={"size": 24}, cbar=False, cmap='Blues', fmt='g')# font size
ax.xaxis.tick_top()
ax.xaxis.set_ticks_position('none')
plt.yticks(va='center', size=16)
plt.xticks(va='center', size=16)
plt.ylabel('True \n', fontsize=20, fontweight='bold')
plt.title('Predicted \n', fontsize=20, fontweight='bold')
plt.xlabel('\n GUESS WHAT?', fontsize=28, fontweight='heavy')
plt.savefig('GWConfusion.eps')

plt.show()


