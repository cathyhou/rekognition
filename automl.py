from openpyxl import load_workbook

wb = load_workbook(filename = 'results.xlsx')
ws = wb['results']

# calculate how many images of each emotion in test image dataset

results = [0, 0, 0, 0, 0]

def toEmotion(emo):
    current = ['HAPPY', 'NEUTRAL', 'SCARED', 'ANGRY', 'SAD']
    for i in range(5):
        if emo == current[i]:
            return i
    return 'error'


for r in range(128):
    row = str(r+2)
    v = ws['E'+row].value
    current = results[toEmotion(v)]
    results[toEmotion(v)] = current + 1

print(results)
